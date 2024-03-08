from  utinities import *
from openpyxl import load_workbook
import pandas as pd

def custom_standardization(df, columns=None):
    # 如果未指定列，使用默认的固定列
    if columns is None:
        columns = ['DC', 'quantity']

    try:
        for column in columns:
            # 检查指定的列是否存在于数据框中
            if column not in df.columns:
                print(f"Warning: Column '{column}' does not exist in the DataFrame.")
                continue

            # 针对指定列进行标准化
            df[column] = pd.to_numeric(df[column], errors='coerce')
            df[column] = df[column].round().astype('Int64')

    except ValueError as e:
        print(f"Error: {e}")

    return df

def read_stock_data(workbook_path, sheet_name):
    workbook = load_workbook(workbook_path, data_only=True)

    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"工作表名稱 '{sheet_name}' 在工作簿中不存在。")

    worksheet = workbook[sheet_name]

    # 使用單一的循環來提取數據值和註釋，並記錄行號
    data = []
    remarks = []
    row_indices = []
    for i, row in enumerate(worksheet['B4':'J' + str(worksheet.max_row)], start=4):
        data.append([cell.value for cell in row])
        remarks.append(row[1].comment.text if row[1].comment else None)
        row_indices.append(i)

    columns = ['supplier_code', 'part_number', 'lot', "date_code", 'DC', 'QTY', 'quantity', 'Y', 'store']
    df = pd.DataFrame(data, columns=columns)

    # 將不需要的列刪除，並加入註釋和行號
    df = df.drop(columns=['supplier_code', 'QTY', 'Y'])
    df['remark'] = remarks
    df['row_index'] = row_indices

    return df


def preprocess_input_dataframe(input_dataframe):
    # 对 input_dataframe 进行加工和分组
    agg_dict = {
        'quantity': 'sum',
        'unit_price': 'first',
        'currency': 'first',
        'invoice_series': 'first',
        'delivery_date': 'first',
        'customer_no': 'first'
    }
    processed_dataframe = input_dataframe.groupby(['part_number', 'product_number', 'customer_part_number', 'purchase_order'], as_index=False).agg(agg_dict)
    processed_dataframe['delivery_date'] = processed_dataframe['delivery_date'].map(format_date_code)
    processed_dataframe['month'] = processed_dataframe['delivery_date'].apply(lambda x: x.split('/')[1])
    processed_dataframe['day'] = processed_dataframe['delivery_date'].apply(lambda x: x.split('/')[2])
    return processed_dataframe






