import pandas as pd
import numpy as np
from openpyxl import load_workbook
import json
from openpyxl.utils import get_column_letter
import os

def is_file_locked(file_path):
    try:
        # 尝试以写入模式打开文件，如果成功说明文件没有被其他进程占用
        with open(file_path, 'a'):
            pass
        return False
    except PermissionError:
        # 如果抛出PermissionError，则文件已经被其他进程打开
        return True

def load_config(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        config = json.load(file)
    return config


def read_excel_data(workbook_path, sheet_name):
    """
    從指定的 Excel 文件和工作表中讀取數據並返回為 DataFrame。

    參數:
        workbook_path (str): Excel 文件的路徑。
        sheet_name (str): 工作表名稱。

    返回:
        pd.DataFrame: 從Excel工作表中讀取的數據。
    """
    # 讀取 Excel 數據到 DataFrame
    df = pd.read_excel(workbook_path, sheet_name=sheet_name)

    # 對 DataFrame 中的每個元素應用 strip 操作
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)

    return df

def clear_worksheet(workbook_path, sheet_name):
    # Load the workbook
    workbook = load_workbook(workbook_path)

    # Get the specific worksheet
    worksheet = workbook[sheet_name]

    # Determine the last row with values
    last_row_with_value = worksheet.max_row

    # Clear the values from row 2 to the last row with values
    for row in worksheet.iter_rows(min_row=2, max_row=last_row_with_value):
        for cell in row:
            cell.value = None

def format_date_code(date_code):
    date_str = str(date_code)
    # 將YYYYMMDD格式的日期碼轉換為YYYY/MM/DD格式
    if len(date_str) == 8:
        date_str = f'{date_str[:4]}/{date_str[4:6]}/{date_str[6:]}'
    parsed_date = pd.to_datetime(date_str, errors='coerce')
    if pd.notna(parsed_date):
        return parsed_date.strftime('%Y/%m/%d')
    else:
        return str(date_code)


# 打開工作簿並檢查是否為唯讀
def open_workbook(file_path):
    try:
        return pd.ExcelFile(file_path)
    except Exception as e:
        print("檔案已被開啟")
        exit()

# 保存工作簿
def save_workbook(file_path, data_range, sheet_name):
    with pd.ExcelWriter(file_path) as writer:
        data_range.to_excel(writer, sheet_name=sheet_name, index=False)

def to_excel_auto_column_width(df: pd.DataFrame, writer: pd.ExcelWriter, sheet_name="Sheet1"):
    """Write DataFrame to Excel with auto-adjusted column widths"""
    try:
        # Write DataFrame to Excel with specified sheet name
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Calculate the character widths for each column header
        column_widths = df.columns.to_series().apply(lambda x: len(str(x).encode('gbk'))).values

        # Calculate the maximum character width for each column
        max_widths = df.astype(str).applymap(lambda x: len(str(x).encode('gbk'))).agg(max).values

        # Take the maximum width for each column from the above two calculations
        widths = np.max([column_widths, max_widths], axis=0)

        # Set the column width for each column in the specified sheet
        worksheet = writer.sheets[sheet_name]
        for i, width in enumerate(widths, 1):
            # Adjusting for openpyxl engine's character width discrepancy
            worksheet.column_dimensions[get_column_letter(i)].width = width + 2

        # Optionally, you can return some information like the number of rows and columns written
        return f"Successfully wrote DataFrame to Excel. Rows: {df.shape[0]}, Columns: {df.shape[1]}"

    except PermissionError as e:
        return f"PermissionError: {e}"
    except Exception as e:
        return f"An error occurred: {e}"


def output_data(workbook_path, sheet_name, dataframe):
    try:
        with pd.ExcelWriter(workbook_path, engine='openpyxl', mode='a') as writer:
            if sheet_name in writer.book.sheetnames:
                writer.book.remove(writer.book[sheet_name])
            to_excel_auto_column_width(dataframe, writer, sheet_name=sheet_name)

    except PermissionError as e:
        print(f"PermissionError: {e}")
        print("The file is currently open. Please close it and try again.")
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        print(f"Dataframe has been successfully output to sheet '{sheet_name}' in the Excel file.")

def ask_deduct_stock():
    """
    Ask the user if they want to deduct stock.

    Returns:
        bool: True if the answer is '1', False otherwise.
    """
    response = input("Do you want to deduct stock? (Enter 1 for yes, 2 for no): ").strip()
    return response == '1'
