import pandas as pd
from openpyxl import load_workbook
import json
import os

def load_config(file_path):
    with open(file_path, 'r') as file:
        config = json.load(file)
    return config

def read_excel_data(workbook_path, sheet_name):
    """
    從指定的 Excel 文件和工作表中讀取數據並返回為 DataFrame。

    參數:
        workbook_path (str): Excel 文件的路徑。
        sheet_name (str): 工作表名稱。

    返回:
        pd.DataFrame: 從Excel工作表中讀取的數據。
    """
    return pd.read_excel(workbook_path, sheet_name=sheet_name)

def clear_worksheet(workbook_path, sheet_name):
    # Load the workbook
    workbook = load_workbook(workbook_path)

    # Get the specific worksheet
    worksheet = workbook[sheet_name]

    # Determine the last row with values
    last_row_with_value = worksheet.max_row

    # Clear the values from row 2 to the last row with values
    for row in worksheet.iter_rows(min_row=2, max_row=last_row_with_value):
        for cell in row:
            cell.value = None

def format_date_code(date_code):
    date_str = str(date_code)
    # 將YYYYMMDD格式的日期碼轉換為YYYY/MM/DD格式
    if len(date_str) == 8:
        date_str = f'{date_str[:4]}/{date_str[4:6]}/{date_str[6:]}'
    parsed_date = pd.to_datetime(date_str, errors='coerce')
    if pd.notna(parsed_date):
        return parsed_date.strftime('%Y/%m/%d')
    else:
        return str(date_code)


# 打開工作簿並檢查是否為唯讀
def open_workbook(file_path):
    try:
        return pd.ExcelFile(file_path)
    except Exception as e:
        print("檔案已被開啟")
        exit()

# 保存工作簿
def save_workbook(file_path, data_range, sheet_name):
    with pd.ExcelWriter(file_path) as writer:
        data_range.to_excel(writer, sheet_name=sheet_name, index=False)

def output_data(workbook_path, sheet_name, dataframe):
    try:
        with pd.ExcelWriter(workbook_path, engine='openpyxl', mode='a') as writer:
            if sheet_name in writer.book.sheetnames:
                writer.book.remove(writer.book[sheet_name])

            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)

    except PermissionError as e:
        print(f"PermissionError: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        print("Ensure that the file is closed.")

def ask_deduct_stock():
    """
    Ask the user if they want to deduct stock.

    Returns:
        bool: True if the answer is '1', False otherwise.
    """
    response = input("Do you want to deduct stock? (Enter 1 for yes, 2 for no): ").strip()
    return response == '1'
